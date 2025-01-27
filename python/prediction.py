import argparse
import pandas as pd
import numpy as np
from sklearn.svm import SVR
from sklearn.ensemble import RandomForestRegressor
from sklearn.tree import DecisionTreeRegressor
from sklearn.linear_model import LinearRegression
from sklearn.linear_model import Lasso, LassoCV
from sklearn.linear_model import Ridge, RidgeCV
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import mean_squared_error, mean_absolute_error
from sklearn.metrics import r2_score
import matplotlib.pyplot as plt
from data_util import *
from evaluation import *

pd.set_option('display.max_rows', None)
pd.set_option('display.max_columns', None)
pd.set_option('display.width', 1000)
pd.set_option('display.colheader_justify', 'left')

parser = argparse.ArgumentParser(description="Agruments")

parser.add_argument('--use_abundant', type=int, default=1, help="use abundant_medal_data or few_medal_data")
parser.add_argument('--years_back', type=int, default=3, help="years to call back")
parser.add_argument('--prediction_year', type=int, default=2024, help="prediction year")
parser.add_argument('--model_type', type=str, default='SVM', help="which model to use")
parser.add_argument('--medal_type', type=str, default='Total', help="which medal to predict")
parser.add_argument('--save', type=int, default=0, help="if save the result")
parser.add_argument('--random_seed', type=int, default=2025, help="random seed")

args = parser.parse_args()

use_abundant = args.use_abundant
years_back = args.years_back
prediction_year = args.prediction_year
model_type = args.model_type
medal_type = args.medal_type
save = args.save
random_seed = args.random_seed

# Data Processing 
match_table = pd.read_csv('../data/match.csv')

medal_data = pd.read_csv('../data/summerOly_medal_counts.csv')
medal_data['NOC'] = medal_data['NOC'].str.strip()
grouped = medal_data.groupby('NOC', group_keys=False)
few_data_indices = grouped.apply(filter_conditions).explode().dropna().astype(int)
few_medal_data = medal_data.loc[few_data_indices]
abundant_medal_data = medal_data.drop(few_data_indices)

#choose_medal_data = abundant_medal_data if use_abundant else few_medal_data
choose_medal_data = medal_data

athletes_data = pd.read_csv('../data/summerOly_athletes.csv')
athletes_data = pd.merge(athletes_data, match_table, left_on='NOC', right_on='abbr', how='left')
athletes_data = athletes_data[['Name', 'Year', 'name']]
athletes_data = athletes_data.rename(columns={
    'name': 'NOC',
    'Name': 'Athletes'
})
athletes_data = athletes_data.groupby(['NOC', 'Year'])['Athletes'].count().reset_index()

host_data = pd.read_csv('../data/summerOly_hosts.csv')
def extract_country(host):
    if "Cancelled" in host:
        return None
    host = host.split('(')[0].strip()
    country = host.split(",")[-1].strip()

    if country == "United Kingdom":
        country = "Great Britain"
    return country
host_data["Country"] = host_data["Host"].apply(extract_country)

events_data = pd.read_csv('../data/summerOly_programs.csv', encoding='Windows-1252')
events_data = events_data.iloc[-3, 4:].reset_index()
events_data.columns = ['Year', 'Events']
events_data = events_data.drop(index=3).reset_index(drop=True)
events_data['Year'] = events_data['Year'].astype(int)

merged_data = pd.merge(medal_data, athletes_data, on=['NOC', 'Year'], how='left')
merged_data = pd.merge(merged_data, host_data, on='Year', how='outer')
merged_data['Is_Host'] = (merged_data['Country'] == merged_data['NOC']).astype(int)
merged_data = pd.merge(merged_data, events_data, on='Year', how='left')

medal_pivot = merged_data.pivot(index='NOC', columns='Year', values=medal_type).fillna(0)
athletes_pivot = merged_data.pivot(index='NOC', columns='Year', values='Athletes').fillna(0)
host_pivot = merged_data.pivot(index='NOC', columns='Year', values='Is_Host').fillna(0)
events_pivot = merged_data.pivot(index='NOC', columns='Year', values='Events')
for col in events_pivot.columns:
    non_missing_values = events_pivot[col][events_pivot[col] != 0].dropna()
    if not non_missing_values.empty:
        fill_value = non_missing_values.iloc[0]
        events_pivot[col] = events_pivot[col].replace(0, fill_value).fillna(fill_value)

years = [1896, 1900, 1904, 1908, 1912, 1920, 1924, 1928, 1932, 1936, 1948, 1952, 1956, 1960, 1964, 1968, 1972, 1976, 1980, 1984, 1988, 1992, 1996, 2000, 2004, 2008, 2012, 2016, 2020]
if prediction_year == 2028:
    years.append(2024)

X = []
y = []

for country, group in choose_medal_data.groupby('NOC'):
    for i in range(len(years) - years_back):
        if years[i + years_back] not in group['Year'].values:
            continue
        medals = [medal_pivot.loc[country, years[i + j]] for j in range(years_back)]
        athletes = [athletes_pivot.loc[country, years[i + years_back]]]
        host = [host_pivot.loc[country, years[i + years_back]]]
        year = [np.float64(years[i + years_back])]
        events = [events_pivot.loc[country, years[i + years_back]]]
        features = medals + athletes + host + year + events
        medal_next = medal_pivot.loc[country, years[i + years_back]]
        X.append(features)
        y.append(medal_next)

X = np.array(X)
y = np.array(y)

noise_factor = 0.1

historical_medal_noise = 1 + np.random.normal(0, noise_factor, X[:, :years_back].shape)
athletes_noise = 1 + np.random.normal(0, noise_factor, X[:, years_back].shape)  
events_noise = 1 + np.random.normal(0, noise_factor, X[:, years_back + 3].shape)  

X_historical_medal_noise = X[:, :years_back].astype(np.float64) * historical_medal_noise
X_athletes_noise = X[:, years_back].astype(np.float64) * athletes_noise
X_events_noise = X[:, years_back + 3].astype(np.float64) * events_noise

X_with_noise_historical = np.hstack([X_historical_medal_noise, X[:, years_back:]])
X_with_noise_athletes = np.hstack([X[:, :years_back], X_athletes_noise.reshape(-1, 1), X[:, years_back+1:]])
X_with_noise_events = np.hstack([X[:, :years_back+3], X_events_noise.reshape(-1, 1), X[:, years_back+4:]])

scaler = StandardScaler()

X_original_scaled = scaler.fit_transform(X)

X_with_noise_historical_scaled = scaler.fit_transform(X_with_noise_historical)
X_with_noise_athletes_scaled = scaler.fit_transform(X_with_noise_athletes)
X_with_noise_events_scaled = scaler.fit_transform(X_with_noise_events)

X_train, X_test, y_train, y_test = train_test_split(X_original_scaled, y, test_size=0.2, random_state=random_seed)
X_train_noise_historical, X_test_noise_historical, y_train_noise_historical, y_test_noise_historical = train_test_split(X_with_noise_historical_scaled, y, test_size=0.2, random_state=random_seed)
X_train_noise_athletes, X_test_noise_athletes, y_train_noise_athletes, y_test_noise_athletes = train_test_split(X_with_noise_athletes_scaled, y, test_size=0.2, random_state=random_seed)
X_train_noise_events, X_test_noise_events, y_train_noise_events, y_test_noise_events = train_test_split(X_with_noise_events_scaled, y, test_size=0.2, random_state=random_seed)

# Model Defination
if model_type == 'SVM':
    model = SVR(kernel='rbf', C=100, gamma=0.1, epsilon=0.1)
elif model_type == 'RandomForest':
    model = RandomForestRegressor(n_estimators=100, random_state=random_seed)
elif model_type == 'DecisionTree':
    model = DecisionTreeRegressor(max_depth=10, random_state=random_seed)
elif model_type == 'LinearRegression':
    model = LinearRegression()
elif model_type == 'Ridge':
    model = Ridge(alpha=1.0)
elif model_type == 'Lasso':
    model = Lasso(alpha=0.1)
elif model_type == 'WeightLinearRegression':
    model = LinearRegression()
elif model_type == 'RidgeCV':
    model = RidgeCV(alphas=(np.linspace(0.1,10.0,num=30)),
                    fit_intercept=True)
elif model_type == 'LassoCV':
    model = LassoCV(alphas=(np.linspace(0.01,10.0,num=100)),
                        fit_intercept=True)
else:
    raise
    
if model_type == 'WeightLinearRegression':
    w = np.exp(-(y_train-50)**2/2000)
    model.fit(X_train, y_train, sample_weight=w)
else:
    model.fit(X_train, y_train)
    y_pred_original = model.predict(X_test)

    model.fit(X_train_noise_historical, y_train_noise_historical)
    y_pred_noise_historical = model.predict(X_test_noise_historical)

    model.fit(X_train_noise_athletes, y_train_noise_athletes)
    y_pred_noise_athletes = model.predict(X_test_noise_athletes)

    model.fit(X_train_noise_events, y_train_noise_events)
    y_pred_noise_events = model.predict(X_test_noise_events)

    plt.figure(figsize=(8, 6))
    plt.scatter(y_pred_original, y_pred_noise_historical, color='red', label='Noisy Historical Medal Data')
    plt.scatter(y_pred_original, y_pred_noise_athletes, color='green', label='Noisy Athletes Data')
    plt.scatter(y_pred_original, y_pred_noise_events, color='blue', label='Noisy Events Data')
    plt.plot([y_test.min(), y_test.max()], [y_test.min(), y_test.max()], color='black', linestyle='--')
    plt.title("Predictions Comparison: Original vs Noisy Data")
    plt.xlabel("Predicted Medal Counts (Original Data)")
    plt.ylabel("Predicted Medal Counts (Noisy Data)")
    plt.legend()
    plt.show()

# Validation
y_pred = model.predict(X_test)
mse = mean_squared_error(y_test, y_pred)
mae = mean_absolute_error(y_test, y_pred)
r2 = r2_score(y_test, y_pred)

#print(f"Test MSE: {mse}")
#print(f"Test MAE: {mae}")
combined_data_path = '../data/clustered_data.csv'

few_medal_df, raw_atheletes_df = preprocess_data(combined_data_path, prediction_year)

medal_winners = evaluate_athletes_by_group(raw_atheletes_df)
medal_winners = medal_winners.merge(match_table, left_on='NOC', right_on='abbr', how='left')
medal_winners = medal_winners[['Year', 'Event', 'Name', 'name', 'Awarded_Medal']]
medal_winners = medal_winners.rename(columns={'name': 'NOC'})

few_medal_df = few_medal_df.merge(match_table, left_on='NOC', right_on='abbr', how='left')
few_medal_df = few_medal_df[['name']]
few_medal_df = few_medal_df.rename(columns={'name': 'NOC'})

medal_prediction = predict_nation_medals(few_medal_df[['NOC']], medal_winners)

actuals = []
predictions = []
choose_medal_data['prediction_result'] = None
for country, group in choose_medal_data.groupby('NOC'):
    if country_filter(medal_data, athletes_data, country, prediction_year) == 2:
        prediction = 'ABSENT'

    elif use_abundant ^ country_filter(medal_data, athletes_data, country, prediction_year):
        input_medals = [medal_pivot.loc[country, years[-years_back + i]] for i in range(years_back)]
        input_host = [host_pivot.loc[country, prediction_year]]
        input_year = [np.float64(prediction_year)]
        if prediction_year == 2028:
            input_events = [predict_2028_events_num(events_data)]
            input_athletes = [predict_2028_athletes_num(athletes_data, country)]
        else:
            input_events = [events_pivot.loc[country, prediction_year]]
            input_athletes = [athletes_pivot.loc[country, prediction_year]]
        input_data = input_medals + input_athletes + input_host + input_year + input_events

        input_data = np.array(input_data).reshape(1, -1)
        input_data = scaler.transform(input_data)

        prediction = model.predict(input_data)[0]
    else: 
        matching_rows = medal_prediction.loc[medal_prediction['NOC'] == country, medal_type]
        if not matching_rows.empty:
            prediction = matching_rows.iloc[0]
        else:
            prediction = 0
    
    if prediction != 'ABSENT':
        prediction = round(prediction)
    choose_medal_data.loc[choose_medal_data['NOC'] == country, 'prediction_result'] = prediction

    if not save:
        print(f'\n{country}')
    if prediction_year == 2028:
        actual_2024 = medal_pivot.loc[country, 2024]
        if not save:
            print(f"{country} - Actual 2024 {medal_type} medal number: ", actual_2024)
            print(f"{country} - Predicted 2028 {medal_type} medal number: ", prediction)
    else:    
        actual = medal_pivot.loc[country, prediction_year]
        print(f"{country} - Actual 2024 {medal_type} medal number: ", actual)
        print(f"{country} - Predicted 2024 {medal_type} medal number: ", prediction)

        if prediction != 'ABSENT':
            predictions.append(prediction)
            actuals.append(actual)

if prediction_year != 2028:
    actuals = np.array(actuals)
    predictions = np.array(predictions)
    plt.figure(figsize=(8, 6))
    plt.scatter(actuals, predictions, color='blue', label='Predicted vs Actual')
    plt.plot([actuals.min(), actuals.max()], [actuals.min(), actuals.max()], color='red', linestyle='--', label="y = x")

    plt.title("Actual vs Predicted Values")
    plt.xlabel("Actual Values")
    plt.ylabel("Predicted Values")

    plt.legend()

    plt.show()
    #print(f'MSE: {mse}')
    #print(f'MAE: {mae}')
    #print(f'R^2: {r2}')
elif save:
    import openpyxl
    import os

    excel_file = f"../result/result_{medal_type}_back{years_back}.xlsx"
    if not os.path.exists(excel_file):
        df = pd.DataFrame()
        df.to_excel(excel_file, index=False)
        medal_pivot_reset = medal_pivot.reset_index()
        noc_column = medal_pivot_reset[['NOC']]
        noc_column.to_excel(excel_file, index=False)

    choose_medal_data_unique = choose_medal_data.drop_duplicates(subset=['NOC'])

    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    existing_columns = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]

    target_column_name = f'{model_type}'
    if target_column_name not in existing_columns:
        predicted_column_index = ws.max_column + 1
        ws.cell(row=1, column=predicted_column_index, value=target_column_name)
    else:
        predicted_column_index = existing_columns.index(target_column_name) + 1

    excel_nocs = [ws.cell(row=row, column=1).value for row in range(2, ws.max_row + 1)]

    for country in choose_medal_data_unique['NOC']:
        if country in excel_nocs:
            excel_row = excel_nocs.index(country) + 2
            prediction = choose_medal_data_unique.loc[choose_medal_data_unique['NOC'] == country, 'prediction_result'].values[0]

            target_cell = ws.cell(row=excel_row, column=predicted_column_index)

            if target_cell.value is None:
                target_cell.value = prediction

    wb.save(excel_file)
